import openpyxl
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import os
from django.conf import settings
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

def register_cyrillic_fonts():
    font_path = '/usr/share/fonts/TTF/DejaVuSans.ttf'
    font_bold = '/usr/share/fonts/TTF/DejaVuSans-Bold.ttf'
    if os.path.exists(font_path) and os.path.exists(font_bold):
        pdfmetrics.registerFont(TTFont('Roboto', font_path))
        pdfmetrics.registerFont(TTFont('Roboto-Bold', font_bold))
        return True
    return False

# For basic cyrillic support without custom fonts we can fallback, 
# but reportlab by default does not support Cyrillic in standard Type1 fonts like Helvetica.
# To ensure our report does not crash, we register a basic font setup if possible, 
# or use UTF-8 strings. Since standard PDFs have encoding issues, we will just pass strings.
# A perfect solution would use `ttfonts` module to load a local `.ttf`, but this works for proof of concept.

def generate_excel_report(data, filename, columns):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num).value = str(cell_value)
            
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response

def generate_pdf_report(data, filename, columns, title):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    
    # Clean data to string
    table_data = [columns] + [[str(cell) for cell in row] for row in data]
    
    t = Table(table_data, repeatRows=1)
    
    has_fonts = register_cyrillic_fonts()
    font_regular = 'Roboto' if has_fonts else 'Helvetica'
    font_bold = 'Roboto-Bold' if has_fonts else 'Helvetica-Bold'
    
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.darkblue),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.lightgrey),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('FONTNAME', (0,0), (-1,0), font_bold),
        ('FONTNAME', (0,1), (-1,-1), font_regular),
    ]))
    
    if title:
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle', 
            parent=styles['Heading1'], 
            fontName=font_bold, 
            fontSize=16, 
            spaceAfter=20
        )
        elements.append(Paragraph(title, title_style))
        
    elements.append(t)
    doc.build(elements)
    
    return response
